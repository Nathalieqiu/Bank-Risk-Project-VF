import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
from openpyxl.drawing.image import Image as XLImage



# PARAMETRES

# Fichier source
INPUT_EXCEL = Path("/Users/nathalyqiu/Desktop/data/raw/bank_database.xlsx")

# Fichier final
OUTPUT_EXCEL = Path("/Users/nathalyqiu/Desktop/data/raw/bank_risk_report.xlsx")

PD_WATCHLIST_THRESHOLD = 0.03 # Seuil = Zone de vigilance utilisé par les banques

# Basel III (proxy)
BASEL_45 = 0.045 # Règlementation = CET1 min
BASEL_8 = 0.08 # Tier 1 + Tier 2 = Capital total min
BASEL_105 = 0.105 # CET1 + Buffer

RW_PROXY = 1.0 # Cas Standard 
CET1_RATE_PROXY = 0.12 # Banque prudente donc sup à 10,5% 

# Dossier pour images 
FIG_DIR = Path("/Users/nathalyqiu/Desktop/data/raw/figures_tmp")


# 1) DONNÉES

if not INPUT_EXCEL.exists():
    raise FileNotFoundError(f"Fichier introuvable : {INPUT_EXCEL}")

clients = pd.read_excel(INPUT_EXCEL, sheet_name="clients")
exposures = pd.read_excel(INPUT_EXCEL, sheet_name="exposures")
collaterals = pd.read_excel(INPUT_EXCEL, sheet_name="collaterals")



# 2) NETTOYAGE

clients = clients.dropna(subset=["client_id"]).copy()
clients["client_id"] = clients["client_id"].astype(str).str.strip()
clients["country"] = clients["country"].astype(str).str.strip()
clients["sector"] = clients["sector"].astype(str).str.strip()
clients["rating"] = clients["rating"].astype(str).str.strip().str.upper()

exposures = exposures.dropna(subset=["client_id", "loan_amount", "ead", "pd", "lgd"]).copy()
exposures["client_id"] = exposures["client_id"].astype(str).str.strip()

exposures["loan_amount"] = pd.to_numeric(exposures["loan_amount"], errors="coerce")
exposures["ead"] = pd.to_numeric(exposures["ead"], errors="coerce")
exposures["pd"] = pd.to_numeric(exposures["pd"], errors="coerce")
exposures["lgd"] = pd.to_numeric(exposures["lgd"], errors="coerce")
exposures = exposures.dropna(subset=["loan_amount", "ead", "pd", "lgd"])

exposures = exposures[(exposures["loan_amount"] > 0) & (exposures["ead"] > 0)]
exposures = exposures[exposures["pd"].between(0, 1)]
exposures = exposures[exposures["lgd"].between(0, 1)]
exposures = exposures[exposures["loan_amount"] <= exposures["ead"]]

collaterals = collaterals.dropna(subset=["client_id", "collateral_value"]).copy()
collaterals["client_id"] = collaterals["client_id"].astype(str).str.strip()
collaterals["collateral_value"] = pd.to_numeric(collaterals["collateral_value"], errors="coerce")
collaterals = collaterals.dropna(subset=["collateral_value"])
collaterals = collaterals[collaterals["collateral_value"] >= 0]



# 3) FUSION

df = exposures.merge(clients, on="client_id", how="left")

coll_agg = collaterals.groupby("client_id", as_index=False).agg(
    total_collateral=("collateral_value", "sum"),
    collateral_lines=("collateral_value", "size"),
)
df = df.merge(coll_agg, on="client_id", how="left")

df["total_collateral"] = df["total_collateral"].fillna(0.0)
df["collateral_lines"] = df["collateral_lines"].fillna(0).astype(int)



# 4) INDICATEURS

df["expected_loss"] = df["ead"] * df["pd"] * df["lgd"]
df["watchlist_flag"] = (df["pd"] > PD_WATCHLIST_THRESHOLD).astype(int)
df["collateral_coverage"] = np.where(df["ead"] > 0, df["total_collateral"] / df["ead"], 0.0)

# Basel III proxy
df["rwa"] = df["ead"] * RW_PROXY
df["cet1_capital"] = df["ead"] * CET1_RATE_PROXY
df["cet1_ratio"] = np.where(df["rwa"] > 0, df["cet1_capital"] / df["rwa"], 0.0)

df["cet1_ok_4_5"] = df["cet1_ratio"] >= BASEL_45
df["cet1_ok_8"] = df["cet1_ratio"] >= BASEL_8
df["cet1_ok_10_5"] = df["cet1_ratio"] >= BASEL_105



# 5) KPIs

kpi_country = df.groupby("country", dropna=False).agg(
    n_clients=("client_id", "count"),
    total_loan=("loan_amount", "sum"),
    total_ead=("ead", "sum"),
    avg_pd=("pd", "mean"),
    total_el=("expected_loss", "sum"),
    watchlist_rate=("watchlist_flag", "mean"),
    avg_collateral_coverage=("collateral_coverage", "mean"),
).reset_index().sort_values("total_ead", ascending=False)

kpi_sector = df.groupby("sector", dropna=False).agg(
    n_clients=("client_id", "count"),
    total_loan=("loan_amount", "sum"),
    total_ead=("ead", "sum"),
    avg_pd=("pd", "mean"),
    total_el=("expected_loss", "sum"),
    watchlist_rate=("watchlist_flag", "mean"),
    avg_collateral_coverage=("collateral_coverage", "mean"),
).reset_index().sort_values("total_ead", ascending=False)

basel_summary = pd.DataFrame({
    "threshold": ["CET1 ≥ 4.5%", "CET1 ≥ 8%", "CET1 ≥ 10.5%"],
    "clients_compliant": [
        int(df["cet1_ok_4_5"].sum()),
        int(df["cet1_ok_8"].sum()),
        int(df["cet1_ok_10_5"].sum())
    ],
    "compliance_rate_%": [
        float(df["cet1_ok_4_5"].mean() * 100),
        float(df["cet1_ok_8"].mean() * 100),
        float(df["cet1_ok_10_5"].mean() * 100)
    ]
})

ead_total = df["ead"].sum()
weighted_pd = (df["pd"] * df["ead"]).sum() / ead_total if ead_total > 0 else 0.0

summary = pd.DataFrame({
    "metric": [
        "Number of clients",
        "Total loan amount",
        "Total EAD",
        "EAD-weighted PD",
        "Total expected loss",
        "WatchList rate",
        "Average collateral coverage"
    ],
    "value": [
        len(df),
        df["loan_amount"].sum(),
        ead_total,
        weighted_pd,
        df["expected_loss"].sum(),
        df["watchlist_flag"].mean(),
        df["collateral_coverage"].mean()
    ]
})


# 6) GRAPHIQUES

FIG_DIR.mkdir(parents=True, exist_ok=True)

pd_plot = FIG_DIR / "pd_distribution.png"
ead_plot = FIG_DIR / "ead_by_country.png"
wl_plot = FIG_DIR / "watchlist_by_sector.png"

plt.figure()
sns.histplot(df["pd"], bins=40, kde=True)
plt.title("PD Distribution")
plt.tight_layout()
plt.savefig(pd_plot)
plt.close()

plt.figure()
sns.barplot(data=kpi_country.head(10), x="country", y="total_ead")
plt.title("Total EAD by Country (Top 10)")
plt.xticks(rotation=30, ha="right")
plt.tight_layout()
plt.savefig(ead_plot)
plt.close()

plt.figure()
sns.barplot(data=kpi_sector.head(10), x="sector", y="watchlist_rate")
plt.title("WatchList Rate by Sector (Top 10 by EAD)")
plt.xticks(rotation=30, ha="right")
plt.tight_layout()
plt.savefig(wl_plot)
plt.close()


# 7) OUTPUT

OUTPUT_EXCEL.parent.mkdir(parents=True, exist_ok=True)

with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="clients_detail", index=False)
    kpi_country.to_excel(writer, sheet_name="kpi_country", index=False)
    kpi_sector.to_excel(writer, sheet_name="kpi_sector", index=False)
    basel_summary.to_excel(writer, sheet_name="basel_compliance", index=False)
    summary.to_excel(writer, sheet_name="summary", index=False)

    # Feuille "graphs"
    wb = writer.book
    ws = wb.create_sheet("graphs")

    ws["A1"] = "PD Distribution"
    ws["A25"] = "Total EAD by Country (Top 10)"
    ws["A49"] = "WatchList Rate by Sector (Top 10 by EAD)"

    img1 = XLImage(str(pd_plot))
    img2 = XLImage(str(ead_plot))
    img3 = XLImage(str(wl_plot))

    # Positionnement des images
    ws.add_image(img1, "A1")
    ws.add_image(img2, "A35")
    ws.add_image(img3, "A70")

print("Fichier Excel final créé :", OUTPUT_EXCEL)

