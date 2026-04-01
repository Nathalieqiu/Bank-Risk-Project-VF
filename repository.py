import json
from pathlib import Path

import pandas as pd
from openpyxl.drawing.image import Image as XLImage

from logger import get_logger

logger = get_logger(__name__)


def load_config(config_path: str = "config.json") -> dict:
 
    with open(config_path, "r") as file:
        config = json.load(file)
    logger.info("Configuration chargée depuis %s", config_path)
    return config


def load_data(input_path: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
   
    path = Path(input_path)

    if not path.exists():
        logger.error("Fichier introuvable : %s", path)
        raise FileNotFoundError(f"Fichier introuvable : {path}")

    clients = pd.read_excel(path, sheet_name="clients")
    exposures = pd.read_excel(path, sheet_name="exposures")
    collaterals = pd.read_excel(path, sheet_name="collaterals")

    logger.info(
        "Données chargées — clients: %d, exposures: %d, collaterals: %d",
        len(clients),
        len(exposures),
        len(collaterals),
    )
    return clients, exposures, collaterals


def save_report(
    output_path: str,
    df: pd.DataFrame,
    kpi_country: pd.DataFrame,
    kpi_sector: pd.DataFrame,
    basel_summary: pd.DataFrame,
    summary: pd.DataFrame,
    figures_dir: str,
) -> None:
    
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    fig_dir = Path(figures_dir)
    pd_plot = fig_dir / "pd_distribution.png"
    ead_plot = fig_dir / "ead_by_country.png"
    wl_plot = fig_dir / "watchlist_by_sector.png"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="clients_detail", index=False)
        kpi_country.to_excel(writer, sheet_name="kpi_country", index=False)
        kpi_sector.to_excel(writer, sheet_name="kpi_sector", index=False)
        basel_summary.to_excel(writer, sheet_name="basel_compliance", index=False)
        summary.to_excel(writer, sheet_name="summary", index=False)

        wb = writer.book
        ws = wb.create_sheet("graphs")

        ws["A1"] = "PD Distribution"
        ws["A35"] = "Total EAD by Country (Top 10)"
        ws["A70"] = "WatchList Rate by Sector (Top 10 by EAD)"

        ws.add_image(XLImage(str(pd_plot)), "A1")
        ws.add_image(XLImage(str(ead_plot)), "A35")
        ws.add_image(XLImage(str(wl_plot)), "A70")

    logger.info("Rapport sauvegardé : %s", path)
